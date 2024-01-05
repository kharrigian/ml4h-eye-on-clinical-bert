
"""
Build Excel workbooks that facilitate span-level note annotation
"""

########################
### Imports
########################

## Standard Library
import re
import os
import sys
import copy
import argparse
import json
from collections import Counter

## External
import numpy as np
import pandas as pd
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

## Local
from cce.util import data_loaders
from cce.util.helpers import chunks
from cce.util.patterns import PHEADER
from cce.util.labels import CONCEPT_ATTRIBUTE_MAP, CONCEPT_BINARY_STATUS

########################
### Globals
#######################

## Borders
HEADER_FMT = Border(top=Side(style="medium"))
THIN_TOP_FMT = Border(top=Side(style="thin"))
THIN_TOP_AND_BOTTOM_FMT = Border(top=Side(style="thin"), bottom=Side(style="thin"))
FOOTER_FMT = Border(bottom=Side(style="medium"))

########################
### Functions
#######################


def parse_command_line():
    """

    """
    ## Initialize Parser
    parser = argparse.ArgumentParser()
    _ = parser.add_argument("--input_file", type=str, default=None, help="Input filepath (CSV or JSON)")
    _ = parser.add_argument("--output_dir", default=None, type=str, help="Output directory")
    _ = parser.add_argument("--output_filename", default=None,  type=str, help="Name of the filename (should end in xlsx)")
    _ = parser.add_argument("--rm_existing", action="store_true", default=False, help="Overwrite existing outfile file.")
    _ = parser.add_argument("--loader_handle_surgical", type=str, default=None, choices={"split","merge"}, help="If included, postprocess surgical span annotations.")
    _ = parser.add_argument("--loader_handle_codes", type=str, default=None, choices={"unseen","all"})
    _ = parser.add_argument("--loader_handle_headers", type=str, default=None, choices={"unseen","all"})
    _ = parser.add_argument("--loader_handle_expand", type=str, default=None, choices={"unseen","all"})
    _ = parser.add_argument("--loader_handle_anti_vegf", action="store_true", default=False, help="If included, try to remove long anti-vegf spans.")    
    _ = parser.add_argument("--loader_sample_rate", default=None, type=float, help="If desired, note-level downsample rate.")
    _ = parser.add_argument("--loader_random_state", default=42, type=int, help="Random seed if doing any note-level downsampling.")
    _ = parser.add_argument("--window_size", type=int, default=10, help="Number of tokens to left/right of span label.")
    _ = parser.add_argument("--sort_by_length", action="store_true", default=False, help="If included, sort documents from shortest to longest.")
    _ = parser.add_argument("--sort_by_examples", action="store_true", default=False, help="If included, sort documents from fewest spans to most spans.")
    _ = parser.add_argument("--batch_size", type=int, default=None, help="Number of notes to include in each batch.")
    _ = parser.add_argument("--batch_inverse_lbl_weight", default=None, type=float, help="Probability of sampling an instance based on inverse label probability. Default is None.")
    _ = parser.add_argument("--batch_uniform_lbl_weight", default=None, type=float, help="Probability of sampling an instance based on uniform label probability. Default is None.")
    _ = parser.add_argument("--n_batches", type=int, default=None, help="Maxium Number of batches to generate.")
    _ = parser.add_argument("--random_state", type=int, default=42, help="Random seed for any sampling.")
    _ = parser.add_argument("--n_jobs", type=int, default=1, help="Number of cores to use for processing.")
    args = parser.parse_args()
    # ## Validate
    if args.output_dir is None or args.output_filename is None:
        raise FileNotFoundError("Must pass an --output_dir and --output_filename")
    if os.path.exists(args.output_dir) and not args.rm_existing:
        raise ValueError("Must include an --rm_existing flag to overwrite existing data files.")
    if not os.path.exists(args.output_dir):
        _ = os.makedirs(args.output_dir)
    if not args.output_filename.endswith(".xlsx"):
        raise ValueError("Output filename should end with .xlsx")
    if args.sort_by_length and args.sort_by_examples:
        raise ValueError("Cannot ask to sort by length AND examples. Must choose one option.")
    ## Cache Arguments
    cfg_suff = args.output_filename.replace(".xlsx",".cfg.json")
    cfg_file = f"{args.output_dir}/{cfg_suff}"
    with open(cfg_file, "w") as the_file:
        _ = json.dump(vars(args), the_file, indent=2)
    ## Return
    return args

def get_window(text,
               span_start,
               span_end,
               window_size=10,
               pheads=None):
    """
    Get a text span and its context window within the text
    """
    ## Check Data
    if text is None or not isinstance(text, str):
        return None, None
    ## Problem Header Match
    phead_span = None
    if pheads is not None:
        phead_span = [p for p in pheads if span_start >= p[0] and span_end <= p[1]]
        phead_span = None if len(phead_span) == 0 else phead_span[0]
    ## Format Based On Whether It's A Problem Header or Not
    if phead_span is not None:
        text_span = text[span_start:span_end]
        text_left = text[phead_span[0]:span_start]
        text_right = text[span_end:phead_span[1]]
        window = f"{text_left} <<{text_span}>> {text_right}"
    else:
        ## Get Spans
        text_left = text[:span_start]
        text_span = text[span_start:span_end]
        text_right = text[span_end:]
        ## Get Window
        window_left = " ".join(text_left.split()[-window_size:])
        window_right = " ".join(text_right.split()[:window_size])
        window = f"{window_left} <<{text_span}>> {window_right}"
    ## Format for Excel Document
    text_span = re.sub(r"[ ]+", " ", text_span.replace("\n"," "))
    window = re.sub(r"[ ]+", " ", window.replace("\n", " "))
    return text_span, window, phead_span is not None

def _format_note_text(text):
    """
    Format note text for slightly easier reading during annotation. Note that this output
    will not align perfectly with span indices due to changes in whitespace.
    """
    ## Early Check
    if text is None or not isinstance(text, str):
        return text
    ## Format
    text = re.sub(r"(\s)(\s+?)(?!\.)", "\n", text[::-1])[::-1]
    text = re.sub(r"\[OVERVIEW\]\s", "[OVERVIEW]\n", text)
    text = re.sub(r"\[ASSESSMENT\s\&\sPLAN\]\s", "[ASSESSMENT & PLAN]\n", text)
    text = re.sub(r"(?<=\n)(\[\[\b)","\n[[", text)
    text = re.sub(r"(?<=\w\n)(\[\[\[PROBLEM LIST\]\]\])", "\n\n[[[PROBLEM LIST]]]", text)
    text = re.sub(r"\]\]\n\[\[\[","]]\n\n[[[", text)
    text = re.sub(r"(?<!\])\]\]\n\n\[\[(?!\[)","]]\n[[", text)
    text = re.sub(r"\[\[\[PROBLEM LIST\]\]\]\n\n\[\[", "[[[PROBLEM LIST]]]\n[[", text)
    text = re.sub(r"\[\[\[ENCOUNTER ICD-10 CODES\]\]\]\n\n\[\[", "[[[ENCOUNTER ICD-10 CODES]]]\n[[", text)
    text = re.sub(r"\n\n(\n+)","\n\n", text)
    return text

def _sample_batches(documents_context,
                    batch_size=None,
                    inverse_weight=None,
                    uniform_weight=None,
                    n_batches=None,
                    random_state=42):
    """

    """
    ## Base Case (No Batches)
    if batch_size is None:
        doc_ids = [sorted(documents_context["document_id"].unique())]
        return doc_ids
    ## Check
    if uniform_weight is not None and inverse_weight is not None:
        raise ValueError("Cannot request inverse_weight and uniform sampling together")
    ## Seed
    seed = np.random.RandomState(random_state)
    ## Count Labels
    lbl_counts = Counter(documents_context["label"])
    lbl_counts_n = sum(lbl_counts.values())
    ## Create a Copy of The Annotations
    documents_context_f = documents_context.reset_index(drop=True).copy()
    doc_id_groups = documents_context_f.groupby(["document_id"]).groups
    ## Document IDs
    lbl2docs = documents_context_f.groupby(["label"])["document_id"].unique().map(set).map(sorted).to_dict()
    all_docs = sorted(documents_context_f["document_id"].unique())
    ## Sample Until All Documents Done
    n_documents = len(documents_context_f["document_id"].unique())
    sampled = set()
    sampled_batches = []
    early_exit = False
    while len(sampled) < n_documents:
        batch = set()
        while len(batch) < batch_size and lbl_counts_n > 0:
            ## Decide Whether to Use Weighting
            if uniform_weight is not None and seed.random() <= uniform_weight:
                ## Identify Remaining Labels
                lbls_avail = [ll for ll, lc in lbl_counts.items() if lc > 0]
                ## Sample Labels Uniformly
                lbl = seed.choice(lbls_avail)
                ## Sample Document
                lbl_doc = seed.choice(lbl2docs[lbl])
            elif inverse_weight is not None and seed.random() <= inverse_weight:
                ## Compute Inverse Sample Weight
                aprops_inv = [(ll, lbl_counts_n / lc) for ll, lc in lbl_counts.items() if lc > 0]
                aprops_norm = sum([i[1] for i in aprops_inv]) 
                aprops_w = [i[1] / aprops_norm for i in aprops_inv]
                ## Sample A Label
                lbl = aprops_inv[seed.choice(len(aprops_w), p=aprops_w)][0]
                ## Sample a Document With That Label
                lbl_doc = seed.choice(lbl2docs[lbl])
            else:
                ## Random Sample
                lbl_doc = seed.choice(all_docs)
            ## Isolate Relevant Rows
            lbl_doc_rows = documents_context_f.iloc[doc_id_groups[lbl_doc]]
            doc_lbl_counts = Counter(lbl_doc_rows["label"].values)
            ## Add Document to Batch
            batch.add(lbl_doc)
            ## Update Counts
            lbl_counts_n -= lbl_doc_rows.shape[0]
            for ll, lc in doc_lbl_counts.items():
                lbl_counts[ll] -= lc
            ## Remove Document From Sample Candidacy
            for ulbl in doc_lbl_counts.keys():
                lbl2docs[ulbl].remove(lbl_doc)
                if len(lbl2docs[ulbl]) == 0:
                    print(f"All {ulbl} instances sampled.")
            all_docs.remove(lbl_doc)
        ## Cache Batch
        sampled_batches.append(list(batch))
        sampled.update(batch)
        ## Update User
        print(">> Batch {:,d} Complete. {:,d} Documents Remaining.".format(len(sampled_batches), n_documents - len(sampled)))
        ## Early Exit
        if n_batches is not None and len(sampled_batches) == n_batches:
            print(">> WARNING: Reached maximum number of batches. Exiting.")
            early_exit = True
            break
    ## Ensure All Documents Added
    if not early_exit:
        assert lbl_counts_n == 0
        assert max(y for x, y in lbl_counts.items()) == 0
    ## Return
    return sampled_batches

def build_annotation_worksheet(documents_context,
                               output_filename):
    """
    
    """
    ## Initialize Workbook
    print("[Initializing Workbook]")
    wb = Workbook()
    _ = wb.remove(wb.active)
    _ = wb.create_sheet(title="annotations")
    ws = wb["annotations"]
    ## Add Data
    row_ind = 0
    for _, doc_id in tqdm(enumerate(documents_context.index.unique()), total=len(documents_context.index.unique()), desc="[Adding Data & Rules]"):
        ## Get Data
        doc_id_df = documents_context.loc[doc_id]
        if isinstance(doc_id_df, pd.Series):
            doc_id_df = doc_id_df.to_frame().T
        ## Get Relevant Subsets of Data
        doc_id_date = "<ENCOUNTER_DATE>"
        doc_id_text = doc_id_df["text"].values[0]
        doc_id_labels = doc_id_df[["text_span_start","text_span_end","label","text_span","text_span_context","Laterality","Severity/Type","Status","Negation","Incorrect Span","Skipped","Notes"]].reset_index(drop=True)
        doc_id_labels = doc_id_labels.sort_values(["text_span_start","text_span_end","label"])
        ## Mapping
        annot_ncol = {col:c for c, col in enumerate(doc_id_labels.columns.tolist())}
        ## Format Data
        row_header = ["document_id", doc_id, *[None for _ in range(doc_id_labels.shape[1] - 2)]]
        row_date = ["encounter_date", doc_id_date, *[None for _ in range(doc_id_labels.shape[1] - 2)]]
        row_text = [doc_id_text, *[None for _ in range(doc_id_labels.shape[1] - 1)]]
        row_labeler = [doc_id_labels.columns.tolist()] + list(map(list, doc_id_labels.values))
        row_empty = [None for _ in range(doc_id_labels.shape[1])]
        ## Add Rows
        ws.append(row_header)
        ws.append(row_date)
        ws.append(row_text)
        for rl in row_labeler:
            ws.append(rl)
        ws.append(row_empty)
        ## Merge Cells
        ws.merge_cells(start_row=row_ind+1, start_column=2, end_row=row_ind+1, end_column=doc_id_labels.shape[1])
        ws.merge_cells(start_row=row_ind+2, start_column=2, end_row=row_ind+2, end_column=doc_id_labels.shape[1])
        ws.merge_cells(start_row=row_ind+3, start_column=1, end_row=row_ind+3, end_column=doc_id_labels.shape[1])
        ws.merge_cells(start_row=row_ind+4+len(row_labeler), start_column=1, end_row=row_ind+4+len(row_labeler), end_column=doc_id_labels.shape[1])
        ## Formatting
        for adder in [1, 2]:
            for c, cell in enumerate(ws[row_ind+adder]):
                if c == 0:
                    font = copy.copy(cell.font)
                    font.b = True
                    cell.font = font
                cell.fill = PatternFill("solid", fgColor="edebeb")
                if adder == 1:
                    cell.border = HEADER_FMT
                elif adder == 2:
                    cell.border = FOOTER_FMT
        alignment = copy.copy(ws[row_ind+3][0].alignment)
        alignment.wrapText=True
        alignment.vertical="top"
        ws[row_ind+3][0].alignment = alignment
        ws.row_dimensions[row_ind+3].height = 200
        for c, cell in enumerate(ws[row_ind+3]):
            cell.border = THIN_TOP_FMT
        for c, cell in enumerate(ws[row_ind+4]):
            font = copy.copy(cell.font)
            font.b = True
            cell.font = font
            cell.border = THIN_TOP_AND_BOTTOM_FMT
        for i, rr in enumerate(range(row_ind + 4 + 1, row_ind + 4 + len(row_labeler))):
            for c, cell in enumerate(ws[rr]):
                if c == annot_ncol["text_span_context"] or c == annot_ncol["Notes"]:
                    alignment = copy.copy(cell.alignment)
                    alignment.wrapText=True
                    alignment.vertical="top"
                    cell.alignment = alignment               
        for rr in range(row_ind + 1, row_ind + 4 + len(row_labeler)):
            for c, cell in enumerate(ws[rr]):
                font = copy.copy(cell.font)
                font.size = "8"
                cell.font = font
        for c, cell in enumerate(ws[rr]):
            cell.border = FOOTER_FMT
        ws.row_dimensions[rr + 1].height = 50
        ## Data Validation
        for rr in range(row_ind + 4 + 1, row_ind + 4 + len(row_labeler)):
            rr_l = None
            rr_lbl = None
            for c, (cell, cell_col) in enumerate(zip(ws[rr], doc_id_labels.columns.tolist())):
                dv_col_id = get_column_letter(c + 1)
                dv = None
                if cell_col == "label":
                    rr_l = cell.value
                    rr_lbl = CONCEPT_ATTRIBUTE_MAP[cell.value.replace(" <<AUTO>>","")]
                if cell_col == "Negation" and rr_l not in CONCEPT_BINARY_STATUS:
                    dv = DataValidation(type="list",formula1='"NEGATED"',allow_blank=True,errorStyle="warning")
                    ws.add_data_validation(dv)
                if cell_col == "Incorrect Span":
                    dv = DataValidation(type="list",formula1='"INCORRECT"',allow_blank=True,errorStyle="warning")
                if rr_lbl is not None and cell_col in rr_lbl:
                    dv = DataValidation(type="list", formula1='"{}"'.format(",".join(rr_lbl[cell_col])), allow_blank=True,errorStyle="warning") 
                if cell_col == "Skipped":
                    dv = DataValidation(type="list",formula1='"SKIPPED"',allow_blank=True,errorStyle="warning")
                if dv is not None:
                    ws.add_data_validation(dv)
                    dv.add(f'{dv_col_id}{rr}:{dv_col_id}{rr}')
        ## Update Row Index
        row_ind += 4 + len(row_labeler)
    ## Column Formatting
    ws.sheet_view.zoomScale = 130
    for col, c in annot_ncol.items():
        col_letter = get_column_letter(c + 1)
        if col == "text_span":
            dim = 20
        elif col == "text_span_context":
            dim = 50
        elif col == "label":
            dim = 20
        elif col == "Notes":
            dim = 50
        else:
            dim = 12
        ws.column_dimensions[col_letter].width = dim
    ## Save and Close
    print("[Caching Annotation File]")
    wb.save(output_filename)
    wb.close()

def main():
    """
    
    """
    ## Parse Command Line
    print("[Parsing Command Line]")
    args = parse_command_line()
    ## Initialize Autolabeler
    print("[Initializing Concept Extractor]")
    autolabeler = data_loaders.AutoLabeler(auto_labels=data_loaders.AUTO_LABELS,
                                           icd10_auto_labels=data_loaders.ICD10_AUTO_LABELS,
                                           handle_icd10_codes=args.loader_handle_codes,
                                           handle_icd10_strings=args.loader_handle_headers,
                                           handle_icd10_expand=args.loader_handle_expand)
    ## Load Inputs
    print("[Loading Documents]")
    documents, _ = data_loaders.load_document_collection(args.input_file,
                                                         autolabeler=autolabeler,
                                                         handle_surgical=args.loader_handle_surgical,
                                                         handle_anti_vegf=args.loader_handle_anti_vegf,
                                                         sample_rate=args.loader_sample_rate,
                                                         random_state=args.loader_random_state,
                                                         verbose=False,
                                                         jobs=args.n_jobs)
    ## Cache Extracted Autolabels
    print("[Caching Autolabels]")
    autolabel_cache_file = "{}/{}".format(args.output_dir, args.output_filename.replace(".xlsx",".source-annotations.json"))
    with open(autolabel_cache_file,"w") as the_file:
        for line in documents:
            the_file.write(f"{json.dumps(line)}\n")
    ## Construct Examples with Context Window
    documents_context = []
    for document in tqdm(documents, total=len(documents), desc="[Formatting Extracted Spans]", file=sys.stdout):
        ## Isolate and Validate Text
        document_text = document["text"]
        if document_text is None or not isinstance(document_text, str):
            continue
        ## Isolate Desired Labels
        document_spans = document["annotations"]
        ## Problem Headers
        document_pheads = [i.span() for i in PHEADER.finditer(document_text)]
        ## Format Text for Output
        document_text_fmt = _format_note_text(document_text)
        ## Iterate Through Labels
        for lbl, lbl_spans in document_spans.items():
            for sstart, send, _ in lbl_spans:
                ## Get Window
                lbl_evidence, lbl_window, lbl_is_phead = get_window(document_text,
                                                                    span_start=sstart,
                                                                    span_end=send,
                                                                    window_size=args.window_size,
                                                                    pheads=document_pheads)
                ## Pre-existing Labels
                notes = None
                laterality = None
                severity_type = None
                status = None
                negation = None
                incorrect = None
                ## Format
                result = {
                    ## Metadata
                    "document_id":document["document_id"],
                    "label":lbl,
                    ## Text Info
                    "text":document_text_fmt,
                    "text_span":lbl_evidence,
                    "text_span_start":sstart,
                    "text_span_end":send,
                    "text_span_context":lbl_window,
                    "text_span_in_header":lbl_is_phead,
                    ## Labels
                    "Laterality":laterality,
                    "Severity/Type":severity_type,
                    "Status":status,
                    "Negation":negation,
                    "Incorrect Span":incorrect,
                    "Skipped":None,
                    ## Extra
                    "Notes":notes
                }
                ## Cache
                documents_context.append(result)
    ## Format
    print("[Initializing Contextualized Document Dataframe]")
    documents_context = pd.DataFrame(documents_context)
    documents_context = documents_context.sort_values(["label","document_id","text_span"],ascending=True)
    documents_context = documents_context.reset_index(drop=True)
    ## Fill Exclusion Cells
    print("[Filling Unnecessary Attribute Cells]")
    for lbl, lbl_types in CONCEPT_ATTRIBUTE_MAP.items():
        lbl_rows = documents_context["label"].str.replace(" <<AUTO>>","") == lbl
        for col in ["Laterality","Severity/Type","Status"]:
            if col not in lbl_types.keys():
                documents_context.loc[lbl_rows, col] = "XXX"
        if lbl in CONCEPT_BINARY_STATUS:
            documents_context.loc[lbl_rows, "Negation"] = "XXX"
    ## Compile Unique Spans
    print("[Caching Span Distribution]")
    lbl2span = documents_context[["label","text_span"]].copy()
    lbl2span["text_span"] = lbl2span["text_span"]
    lbl2span = lbl2span.groupby(["label","text_span"]).size()
    lbl2span = lbl2span.to_frame("count").sort_values("count",ascending=False)
    _ = lbl2span.to_csv("{}/{}".format(args.output_dir, args.output_filename.replace(".xlsx",".label-distribution.csv")))
    print("[Annotations Include a Total of {:,d} Spans]".format(documents_context.shape[0]))
    ## Initialize Batches
    print("[Sampling Batches]")
    doc_batches = _sample_batches(documents_context,
                                  batch_size=args.batch_size,
                                  inverse_weight=args.batch_inverse_lbl_weight,
                                  uniform_weight=args.batch_uniform_lbl_weight,
                                  random_state=args.random_state,
                                  n_batches=args.n_batches)
    ## Initialize file Directory
    print("[Initializing Output Directories for Annotations]")
    if not os.path.exists(f"{args.output_dir}"):
        _ = os.makedirs(f"{args.output_dir}")
   ## Iterate Through Batches
    print("[Generating Batch Files]")
    for d, docs in enumerate(doc_batches):
        ## Select and Sort
        print("[Selecting Batch {}/{}]".format(d+1, len(doc_batches)))
        docs_annotations_context = documents_context.loc[documents_context["document_id"].isin(set(docs))]
        docs_annotations_context = docs_annotations_context.set_index("document_id").sort_index()
        ## Sort
        print("[Sorting Batch]")
        if args.sort_by_length:
            doc_n = docs_annotations_context["text"].drop_duplicates().map(len).to_dict()
            doc_ind = sorted(doc_n.keys(), key=lambda x: doc_n[x])
            docs_annotations_context = docs_annotations_context.loc[doc_ind]
        if args.sort_by_examples:
            doc_n = pd.Series(docs_annotations_context.index).value_counts().to_dict()
            doc_ind = sorted(doc_n.keys(), key=lambda x: doc_n[x])
            docs_annotations_context = docs_annotations_context.loc[doc_ind]
        ## Build Worksheet
        print("[Generating Annotation Worksheet]")
        chunk_output_filename = args.output_filename.replace(".xlsx",f".batch-{d}.xlsx")
        _ = build_annotation_worksheet(documents_context=docs_annotations_context,
                                       output_filename=f"{args.output_dir}/{chunk_output_filename}")
        ## Validate Data Dump
        print("[Validating Saved Data File]")
        df_saved = data_loaders.load_annotated_worksheet(filename=f"{args.output_dir}/{chunk_output_filename}")
        assert df_saved.shape[0] == docs_annotations_context.reset_index().shape[0]
        assert set(df_saved["document_id"]) == set(docs_annotations_context.index)
        assert (df_saved.groupby(["document_id","label"]).size() == docs_annotations_context.reset_index().groupby(["document_id","label"]).size()).all()
        print("[Data Validated. {} Documents ({} Spans) Included]".format(len(df_saved["document_id"].unique()), df_saved.shape[0]))
        ## Batch Maximums
        if args.n_batches is not None and (d + 1) == args.n_batches:
            print("[Maximum Number of Batches Reached ({:,d})]".format(args.n_batches))
            break
    ## Done
    print("[Script Complete]")

########################
### Execution
########################
    
if __name__ == "__main__":
    _ = main()